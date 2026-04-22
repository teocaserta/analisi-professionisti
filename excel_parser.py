"""
excel_parser.py — Parser robusto per due formati contabili:
  1. Prospetto Reddito  (colonne: Incassi, Pagamenti, Reddito)
  2. Conto Economico    (colonne: DES_TIPO_SEZ, Saldo finale, Importo reddito)
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any
import openpyxl


# ─────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────

def parse_excel(file_path: str | Path) -> dict[str, Any]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File non trovato: {file_path}")

    client_id = _extract_client_id(file_path.stem)
    anno      = _extract_year(file_path.stem)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = _find_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Foglio Excel vuoto.")

    header_idx = _find_header_row(rows)
    headers    = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # ── Rilevamento automatico del tipo di file ──
    if any("des_tipo_sez" in h for h in headers):
        file_type = "conto_economico"
        parsed = _parse_conto_economico(rows, header_idx, headers)
    elif any(h in ("incassi", "pagamenti") for h in headers):
        file_type = "prospetto_reddito"
        parsed = _parse_prospetto_reddito(rows, header_idx, headers)
    else:
        raise ValueError(
            "Formato file non riconosciuto. "
            "Attesi: 'Prospetto Reddito' (colonne Incassi/Pagamenti) "
            "o 'Conto Economico' (colonna DES_TIPO_SEZ)."
        )

    ricavi = parsed["ricavi"]
    costi  = parsed["costi"]
    margine = ricavi - costi
    margine_percentuale = (margine / ricavi * 100) if ricavi else 0.0

    return {
        "client_id":           client_id,
        "periodo":             anno,
        "file_type":           file_type,
        "ricavi":              round(ricavi, 2),
        "costi":               round(costi, 2),
        "margine":             round(margine, 2),
        "margine_percentuale": round(margine_percentuale, 2),
        "raw_accounts":        parsed["accounts"],
    }


# ─────────────────────────────────────────────────────────────
# Parser 1 — Prospetto Reddito
# ─────────────────────────────────────────────────────────────

def _parse_prospetto_reddito(rows, header_idx, headers) -> dict:
    col = _map_cols(headers, {
        "descrizione": ["rag.fiscale", "conto", "partitario", "descrizione"],
        "incassi":     ["incassi"],
        "pagamenti":   ["pagamenti"],
        "rettifiche":  ["rettifiche"],
        "reddito":     ["reddito rettificato", "reddito"],
    })
    _require(col, ["descrizione", "incassi", "pagamenti"])

    ricavi_tot = 0.0
    costi_tot  = 0.0
    accounts   = []

    for row in rows[header_idx + 1:]:
        desc = row[col["descrizione"]]
        if desc is None:
            continue
        desc_str = str(desc)
        if not _is_top_level_prospetto(desc_str):
            continue

        incassi    = _num(row[col["incassi"]])
        pagamenti  = _num(row[col["pagamenti"]])
        rettifiche = _num(row[col.get("rettifiche", -1)]) if col.get("rettifiche") is not None else 0.0
        reddito    = _num(row[col.get("reddito", -1)])    if col.get("reddito") is not None else None
        codice     = _extract_code_prospetto(desc_str)

        if incassi > 0:
            ricavi_tot += incassi
            tipo = "ricavo"
        elif pagamenti > 0 or abs(rettifiche) > 0:
            costi_tot += pagamenti + abs(rettifiche)
            tipo = "costo"
        else:
            tipo = "altro"

        accounts.append({
            "codice":              codice,
            "descrizione":         desc_str.strip(),
            "tipo":                tipo,
            "incassi":             incassi,
            "pagamenti":           pagamenti,
            "rettifiche":          rettifiche,
            "reddito_rettificato": reddito,
        })

    return {"ricavi": ricavi_tot, "costi": costi_tot, "accounts": accounts}


def _is_top_level_prospetto(text: str) -> bool:
    return len(text) - len(text.lstrip()) == 0

def _extract_code_prospetto(text: str) -> str | None:
    m = re.match(r"^\s*(\d{2,4})\s*[-–]", str(text))
    return m.group(1) if m else None


# ─────────────────────────────────────────────────────────────
# Parser 2 — Conto Economico
# ─────────────────────────────────────────────────────────────

def _parse_conto_economico(rows, header_idx, headers) -> dict:
    col = _map_cols(headers, {
        "sezione":     ["des_tipo_sez"],
        "conto":       ["conto"],
        "descrizione": ["descrizione"],
        "saldo":       ["saldo finale"],
        "quota_inded": ["quota inded", "quota inded./ non impon."],
        "importo_reddito": ["importo reddito"],
    })
    _require(col, ["sezione", "saldo"])

    ricavi_tot = 0.0
    costi_tot  = 0.0
    accounts   = []

    for row in rows[header_idx + 1:]:
        sezione = str(row[col["sezione"]] or "").strip()
        conto   = str(row[col["conto"]]   or "").strip() if col.get("conto") is not None else ""
        desc    = str(row[col.get("descrizione", col["saldo"])] or "").strip() if col.get("descrizione") is not None else ""
        saldo   = _num(row[col["saldo"]])
        importo_reddito = _num(row[col["importo_reddito"]]) if col.get("importo_reddito") is not None else saldo

        if not sezione or saldo == 0.0:
            continue

        # Solo conti top-level: codice senza punti (es. "47", "55")
        if not _is_top_level_ce(conto):
            continue

        is_ricavo = "ricavi" in sezione.lower()
        is_costo  = "costi"  in sezione.lower()

        if not is_ricavo and not is_costo:
            continue

        # Per i ricavi usa saldo; per i costi usa importo_reddito (già al netto indeducibili)
        if is_ricavo:
            valore = saldo
            ricavi_tot += valore
            tipo = "ricavo"
            pagamenti_out = 0.0
            incassi_out   = valore
        else:
            valore = importo_reddito if importo_reddito else saldo
            costi_tot += valore
            tipo = "costo"
            pagamenti_out = valore
            incassi_out   = 0.0

        accounts.append({
            "codice":              conto,
            "descrizione":         desc,
            "tipo":                tipo,
            "incassi":             incassi_out,
            "pagamenti":           pagamenti_out,
            "rettifiche":          0.0,
            "reddito_rettificato": importo_reddito if is_costo else saldo,
        })

    return {"ricavi": ricavi_tot, "costi": costi_tot, "accounts": accounts}


def _is_top_level_ce(conto: str) -> bool:
    """Top level = codice numerico senza punti (es. '47', '55', '61')."""
    return bool(re.match(r"^\d{2,3}$", conto.strip()))


# ─────────────────────────────────────────────────────────────
# Helpers comuni
# ─────────────────────────────────────────────────────────────

def _extract_client_id(stem: str) -> str:
    m = re.match(r"^(\d+)", stem)
    return m.group(1) if m else "UNKNOWN"

def _extract_year(stem: str) -> str:
    m = re.search(r"(20\d{2})", stem)
    return m.group(1) if m else "UNKNOWN"

def _find_sheet(wb: openpyxl.Workbook):
    preferred = ["foglio1", "sheet1", "prospetto", "reddito", "conto", "economico"]
    for name in wb.sheetnames:
        if name.lower() in preferred:
            return wb[name]
    return wb.active

def _find_header_row(rows: list) -> int:
    signals = {"incassi", "pagamenti", "rettifiche", "descrizione", "conto",
               "reddito", "saldo", "des_tipo_sez", "saldo finale"}
    for i, row in enumerate(rows):
        cells = {str(c).strip().lower() for c in row if c is not None}
        if cells & signals:
            return i
    return 0

def _map_cols(headers: list[str], mapping: dict[str, list[str]]) -> dict[str, int | None]:
    result = {}
    for key, aliases in mapping.items():
        result[key] = None
        for i, h in enumerate(headers):
            if any(alias in h for alias in aliases):
                result[key] = i
                break
    return result

def _require(col_map: dict, required: list[str]) -> None:
    missing = [c for c in required if col_map.get(c) is None]
    if missing:
        raise ValueError(f"Colonne mancanti nel file: {missing}")

def _num(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0
